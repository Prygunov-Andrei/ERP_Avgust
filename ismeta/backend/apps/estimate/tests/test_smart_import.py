"""Тесты E7.2: Smart Excel import — auto-detect, single price, skip totals."""

import io
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font

from apps.estimate.excel.importer import import_estimate_xlsx
from apps.estimate.models import Estimate, EstimateItem, EstimateSection
from apps.workspace.models import Workspace

User = get_user_model()


def _make_xlsx(headers, rows, bold_rows=None):
    bold_rows = bold_rows or set()
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for idx, row in enumerate(rows):
        ws.append(row)
        if idx in bold_rows:
            for c in ws[ws.max_row]:
                c.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture()
def ws():
    return Workspace.objects.create(name="WS-Smart", slug="ws-smart")


@pytest.fixture()
def user():
    return User.objects.create_user(username="smart-user", password="pass")


@pytest.fixture()
def estimate(ws, user):
    return Estimate.objects.create(
        workspace=ws, name="Smart import",
        default_material_markup={"type": "percent", "value": 30},
        default_work_markup={"type": "percent", "value": 300},
        created_by=user,
    )


@pytest.mark.django_db
class TestAutoDetectHeaders:
    def test_standard_headers(self, estimate, ws):
        xlsx = _make_xlsx(
            ["Наименование", "Ед.изм.", "Кол-во", "Цена оборуд.", "Цена мат.", "Цена работ"],
            [["Кабель UTP", "м", 100, 0, 150, 50]],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1

    def test_alternative_headers(self, estimate, ws):
        xlsx = _make_xlsx(
            ["Название", "Единица", "Количество", "Оборудование", "Материалы", "Работы"],
            [["Вентилятор", "шт", 2, 85000, 0, 12000]],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        item = EstimateItem.objects.filter(estimate=estimate, workspace_id=ws.id).first()
        assert item.name == "Вентилятор"
        assert item.equipment_price == Decimal("85000.00")

    def test_short_headers(self, estimate, ws):
        """Ед., К-во, Цена."""
        xlsx = _make_xlsx(
            ["Позиция", "Ед.", "К-во", "Цена"],
            [["Датчик дыма", "шт", 18, 350]],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        item = EstimateItem.objects.filter(estimate=estimate, workspace_id=ws.id).first()
        assert item.material_price == Decimal("350.00")

    def test_no_headers_fallback(self, estimate, ws):
        """Без узнаваемых заголовков → позиционный fallback."""
        xlsx = _make_xlsx(
            ["A", "B", "C", "D", "E", "F"],
            [["Кабель", "м", 50, 0, 100, 50]],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1


@pytest.mark.django_db
class TestSinglePriceColumn:
    def test_one_price_goes_to_material(self, estimate, ws):
        """Единая «Цена» → material_price."""
        xlsx = _make_xlsx(
            ["Наименование", "Ед.изм.", "Кол-во", "Цена"],
            [
                ["Кабель UTP", "м", 100, 150],
                ["Вентилятор", "шт", 2, 85000],
            ],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        items = list(EstimateItem.objects.filter(estimate=estimate, workspace_id=ws.id).order_by("sort_order"))
        assert items[0].material_price == Decimal("150.00")
        assert items[1].material_price == Decimal("85000.00")


@pytest.mark.django_db
class TestSkipTotals:
    def test_skip_itogo_rows(self, estimate, ws):
        xlsx = _make_xlsx(
            ["Наименование", "Ед.изм.", "Кол-во", "Цена"],
            [
                ["Кабель UTP", "м", 100, 150],
                ["Итого", None, None, 15000],
                ["Вентилятор", "шт", 2, 85000],
                ["ВСЕГО", None, None, 185000],
                ["Итого по разделу", None, None, 100000],
            ],
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        assert result.skipped == 3

    def test_skip_total_as_section(self, estimate, ws):
        """Жирная строка 'Итого' не должна стать секцией."""
        xlsx = _make_xlsx(
            ["Наименование", "Ед.изм.", "Кол-во", "Цена"],
            [
                ["Вентиляция", None, None, None],
                ["Кабель", "м", 50, 100],
                ["Итого", None, None, None],  # skip, не секция
                ["Слаботочка", None, None, None],
                ["Датчик", "шт", 10, 350],
            ],
            bold_rows={0, 2, 3},
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        assert result.skipped == 1
        sections = EstimateSection.objects.filter(estimate=estimate)
        names = set(sections.values_list("name", flat=True))
        assert "Вентиляция" in names
        assert "Слаботочка" in names
        assert "Итого" not in names
