"""Тесты Excel import (E7)."""

import io
import uuid
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework.test import APIClient

from apps.estimate.excel.importer import import_estimate_xlsx
from apps.estimate.models import Estimate, EstimateItem, EstimateSection
from apps.estimate.services.estimate_service import EstimateService
from apps.workspace.models import Workspace

User = get_user_model()
WS_HEADER = "HTTP_X_WORKSPACE_ID"


def _make_xlsx(rows, bold_rows=None):
    """Создать .xlsx в памяти. rows — list of lists. bold_rows — set of 0-based indices."""
    bold_rows = bold_rows or set()
    wb = Workbook()
    ws = wb.active
    ws.append(["Наименование", "Ед.изм.", "Кол-во", "Цена оборуд.", "Цена мат.", "Цена работ", "row_id"])
    for idx, row in enumerate(rows):
        ws.append(row)
        if idx in bold_rows:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.fixture()
def ws():
    return Workspace.objects.create(name="WS-Import", slug="ws-import")


@pytest.fixture()
def user():
    return User.objects.create_user(username="import-user", password="pass")


@pytest.fixture()
def client(user):
    c = APIClient()
    c.force_authenticate(user=user)
    return c


@pytest.fixture()
def estimate(ws, user):
    return Estimate.objects.create(
        workspace=ws, name="Import test",
        default_material_markup={"type": "percent", "value": 30},
        default_work_markup={"type": "percent", "value": 300},
        created_by=user,
    )


@pytest.mark.django_db
class TestExcelImport:
    def test_import_new_items(self, estimate, ws):
        xlsx = _make_xlsx([
            ["Кабель UTP", "м", 100, 0, 150, 50, None],
            ["Вентилятор", "шт", 2, 85000, 0, 12000, None],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        assert result.updated == 0
        assert result.errors == []
        assert EstimateItem.objects.filter(estimate=estimate, workspace_id=ws.id).count() == 2

    def test_import_with_row_id_update(self, estimate, ws):
        section = EstimateSection.objects.create(
            estimate=estimate, workspace=ws, name="Тест", sort_order=1,
        )
        item = EstimateService.create_item(section, estimate, ws.id, {
            "name": "Кабель UTP", "unit": "м", "quantity": 50, "material_price": 100,
        })
        old_row_id = str(item.row_id)

        xlsx = _make_xlsx([
            ["Кабель UTP обновлённый", "м", 200, 0, 200, 50, old_row_id],
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.updated == 1
        assert result.created == 0

        updated = EstimateItem.all_objects.get(id=item.id)
        assert updated.name == "Кабель UTP обновлённый"
        assert updated.quantity == Decimal("200.0000")

    def test_import_with_sections(self, estimate, ws):
        xlsx = _make_xlsx(
            [
                ["Вентиляция", None, None, None, None, None, None],
                ["Воздуховод", "м.п.", 42, 0, 800, 200, None],
                ["Слаботочка", None, None, None, None, None, None],
                ["Кабель", "м", 100, 0, 150, 50, None],
            ],
            bold_rows={0, 2},
        )
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 2
        sections = EstimateSection.objects.filter(estimate=estimate)
        assert sections.count() == 2
        assert set(sections.values_list("name", flat=True)) == {"Вентиляция", "Слаботочка"}

    def test_import_empty_file(self, estimate, ws):
        xlsx = _make_xlsx([])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 0
        assert result.updated == 0

    def test_import_invalid_format(self, estimate, ws):
        bad_file = io.BytesIO(b"not an xlsx")
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), bad_file)
        assert result.created == 0
        assert len(result.errors) > 0
        assert "Невалидный" in result.errors[0]

    def test_import_skips_bad_rows(self, estimate, ws):
        xlsx = _make_xlsx([
            [None, "м", 10, 0, 100, 50, None],  # пустое имя
            ["Кабель", "м", -5, 0, 100, 50, None],  # отрицательное кол-во
            ["Хороший кабель", "м", 10, 0, 100, 50, None],  # нормальный
        ])
        result = import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        assert result.created == 1
        assert len(result.errors) == 2

    def test_import_recalcs_totals(self, estimate, ws):
        xlsx = _make_xlsx([
            ["Item 1", "шт", 10, 100, 200, 50, None],
            ["Item 2", "шт", 5, 50, 100, 30, None],
        ])
        import_estimate_xlsx(str(estimate.id), str(ws.id), xlsx)
        estimate.refresh_from_db()
        assert estimate.total_amount > 0

    def test_multipart_upload_api(self, client, estimate, ws):
        xlsx = _make_xlsx([
            ["Кабель API", "м", 50, 0, 150, 50, None],
        ])
        resp = client.post(
            f"/api/v1/estimates/{estimate.id}/import/excel/",
            {"file": SimpleUploadedFile("test.xlsx", xlsx.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            format="multipart",
            **{WS_HEADER: str(ws.id)},
        )
        assert resp.status_code == 200
        assert resp.data["created"] == 1
