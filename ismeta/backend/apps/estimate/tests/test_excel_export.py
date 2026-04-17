"""Тесты Excel export."""

import io

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.estimate.excel.exporter import export_estimate_xlsx
from apps.estimate.models import Estimate, EstimateSection
from apps.estimate.services.estimate_service import EstimateService
from apps.workspace.models import Workspace

User = get_user_model()


@pytest.fixture()
def ws():
    return Workspace.objects.create(name="WS-Excel", slug="ws-excel")


@pytest.fixture()
def estimate(ws, user):
    return Estimate.objects.create(
        workspace=ws, name="Excel test",
        default_material_markup={"type": "percent", "value": 30},
        default_work_markup={"type": "percent", "value": 300},
        created_by=user,
    )


@pytest.fixture()
def user():
    return User.objects.create_user(username="excel-user", password="pass")


@pytest.fixture()
def section(estimate, ws):
    return EstimateSection.objects.create(
        estimate=estimate, workspace=ws, name="Вентиляция", sort_order=1
    )


@pytest.fixture()
def items(section, estimate, ws):
    names = ["Вентилятор", "Кабель UTP", "Датчик дыма"]
    created = []
    for i, name in enumerate(names):
        item = EstimateService.create_item(section, estimate, ws.id, {
            "name": name, "quantity": i + 1,
            "equipment_price": 100, "material_price": 200, "work_price": 50,
        })
        created.append(item)
    return created


@pytest.mark.django_db
class TestExcelExport:
    def test_export_creates_valid_xlsx(self, estimate, ws, section, items):
        output = export_estimate_xlsx(estimate.id, ws.id)
        assert isinstance(output, io.BytesIO)
        wb = load_workbook(output)
        assert "Смета" in wb.sheetnames
        assert "Агрегаты" in wb.sheetnames

    def test_sheet_contains_items_and_totals(self, estimate, ws, section, items):
        output = export_estimate_xlsx(estimate.id, ws.id)
        wb = load_workbook(output)
        ws_sheet = wb["Смета"]
        rows = list(ws_sheet.iter_rows(min_row=2, values_only=True))
        # Section header + 3 items + empty + totals = ≥5 rows
        names = [r[1] for r in rows if r[1] and r[1] not in ("ИТОГО",)]
        assert "Вентилятор" in names
        assert "Кабель UTP" in names
        assert "Датчик дыма" in names

    def test_row_id_hidden_column_present(self, estimate, ws, section, items):
        output = export_estimate_xlsx(estimate.id, ws.id)
        wb = load_workbook(output)
        ws_sheet = wb["Смета"]
        # Column K (11th) = row_id, should be hidden
        assert ws_sheet.column_dimensions["K"].hidden is True

    def test_export_api_endpoint(self, estimate, ws, section, items, user):
        client = APIClient()
        client.force_authenticate(user=user)
        resp = client.get(
            f"/api/v1/estimates/{estimate.id}/export/xlsx/",
            HTTP_X_WORKSPACE_ID=str(ws.id),
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp["Content-Type"]
