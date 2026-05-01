"""Генерация XLSX из result_json.items — публичный download endpoint.

Schema items соответствует recognition.SpecItem (см. recognition/app/schemas/spec.py).
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# (заголовок, ключ в SpecItem)
COLUMNS: list[tuple[str, str]] = [
    ("№ поз", "sort_order"),
    ("Раздел", "section_name"),
    ("Наименование", "name"),
    ("Модель", "model_name"),
    ("Бренд", "brand"),
    ("Производитель", "manufacturer"),
    ("Ед. изм.", "unit"),
    ("Кол-во", "quantity"),
    ("Страница", "page_number"),
]


def _cell_value(item: dict[str, Any], key: str) -> Any:
    value = item.get(key, "")
    if value is None:
        return ""
    return value


def build_workbook(items: Iterable[dict[str, Any]]) -> bytes:
    """Сконструировать .xlsx бинарь из списка items.

    Возвращает байты — для FileResponse / HttpResponse.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    header_font = Font(bold=True)
    for col_idx, (title, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    column_widths = [len(title) for title, _ in COLUMNS]

    for row_idx, item in enumerate(items, start=2):
        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            value = _cell_value(item, key)
            ws.cell(row=row_idx, column=col_idx, value=value)
            length = len(str(value))
            if length > column_widths[col_idx - 1]:
                column_widths[col_idx - 1] = length

    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 8), 80)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
