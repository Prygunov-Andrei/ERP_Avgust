from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from django.db.models import F, ExpressionWrapper, DecimalField, Sum
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone

from core.version_mixin import VersioningMixin
from .models import (
    Project, ProjectNote, Estimate, EstimateSection,
    EstimateSubsection, EstimateCharacteristic, EstimateItem,
    MountingEstimate, ColumnConfigTemplate,
)
from .serializers import (
    ProjectSerializer, ProjectListSerializer, ProjectNoteSerializer,
    EstimateSerializer, EstimateCreateSerializer,
    EstimateSectionSerializer, EstimateSubsectionSerializer,
    EstimateCharacteristicSerializer,
    EstimateItemSerializer, EstimateItemBulkCreateSerializer,
    MountingEstimateSerializer, MountingEstimateCreateFromEstimateSerializer,
    ColumnConfigTemplateSerializer,
)


class ProjectViewSet(VersioningMixin, viewsets.ModelViewSet):
    """ViewSet для проектов (с поддержкой версионирования через VersioningMixin)"""
    
    queryset = Project.objects.select_related(
        'object', 'primary_check_by', 'secondary_check_by', 'parent_version'
    ).prefetch_related('project_notes', 'project_notes__author')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [
        'object', 'stage', 'is_approved_for_production',
        'primary_check_done', 'secondary_check_done'
    ]
    search_fields = ['cipher', 'name']
    version_list_serializer_class = ProjectListSerializer
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ProjectListSerializer
        return ProjectSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # По умолчанию показываем только актуальные версии
        if 'is_current' not in self.request.query_params:
            queryset = queryset.filter(is_current=True)
        return queryset
    
    # Методы versions() и create_version() наследуются от VersioningMixin
    
    @action(detail=True, methods=['post'], url_path='primary-check')
    def primary_check(self, request, pk=None):
        """Отметить первичную проверку"""
        project = self.get_object()
        project.primary_check_done = True
        project.primary_check_by = request.user
        project.primary_check_date = timezone.now().date()
        project.save()
        serializer = ProjectSerializer(project)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='secondary-check')
    def secondary_check(self, request, pk=None):
        """Отметить вторичную проверку"""
        project = self.get_object()
        project.secondary_check_done = True
        project.secondary_check_by = request.user
        project.secondary_check_date = timezone.now().date()
        project.save()
        serializer = ProjectSerializer(project)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'], url_path='approve-production')
    def approve_production(self, request, pk=None):
        """Разрешить "В производство работ" """
        project = self.get_object()
        project.is_approved_for_production = True
        project.production_approval_date = timezone.now().date()
        if 'production_approval_file' in request.FILES:
            project.production_approval_file = request.FILES['production_approval_file']
        project.save()
        serializer = ProjectSerializer(project)
        return Response(serializer.data)


class ProjectNoteViewSet(viewsets.ModelViewSet):
    """ViewSet для замечаний к проектам"""
    
    queryset = ProjectNote.objects.select_related('project', 'author')
    serializer_class = ProjectNoteSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['project']
    
    def perform_create(self, serializer):
        serializer.save(author=self.request.user)


class EstimateViewSet(VersioningMixin, viewsets.ModelViewSet):
    """ViewSet для смет (с поддержкой версионирования через VersioningMixin)"""

    queryset = Estimate.objects.select_related(
        'object', 'legal_entity', 'price_list', 'created_by',
        'checked_by', 'approved_by', 'parent_version'
    ).prefetch_related(
        'projects',
        'sections',
        'sections__subsections',
        'characteristics'
    )
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [
        'object', 'legal_entity', 'status', 'approved_by_customer'
    ]
    search_fields = ['number', 'name']

    def get_serializer_class(self):
        if self.action == 'create':
            return EstimateCreateSerializer
        return EstimateSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        # Аннотируем агрегаты для избежания N+1 в сериализаторе
        if self.action in ('list', 'retrieve'):
            queryset = queryset.annotate(
                _total_materials_sale=Sum('sections__subsections__materials_sale'),
                _total_works_sale=Sum('sections__subsections__works_sale'),
                _total_materials_purchase=Sum('sections__subsections__materials_purchase'),
                _total_works_purchase=Sum('sections__subsections__works_purchase'),
            )
        return queryset
    
    # Методы versions() и create_version() наследуются от VersioningMixin
    
    @action(detail=True, methods=['post'], url_path='create-mounting-estimate')
    def create_mounting_estimate(self, request, pk=None):
        """Создать монтажную смету из обычной сметы"""
        estimate = self.get_object()
        created_by = request.user
        mounting_estimate = MountingEstimate.create_from_estimate(estimate, created_by)
        serializer = MountingEstimateSerializer(mounting_estimate)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], url_path='export')
    def export(self, request, pk=None):
        """Экспорт сметы в Excel с учётом column_config."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from decimal import Decimal

        from .column_defaults import DEFAULT_COLUMN_CONFIG
        from .formula_engine import compute_all_formulas

        estimate = self.get_object()
        config = estimate.column_config or DEFAULT_COLUMN_CONFIG
        visible_cols = [c for c in config if c.get('visible', True)]

        items = EstimateItem.objects.filter(
            estimate=estimate,
        ).select_related('section').order_by('section__sort_order', 'sort_order', 'item_number')

        sections = EstimateSection.objects.filter(estimate=estimate).order_by('sort_order')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Смета'

        bold = Font(bold=True)
        header_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=11, color='1F4E79')
        center = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        num_fmt = '#,##0.00'

        # Title
        num_cols = len(visible_cols)
        last_col_letter = openpyxl.utils.get_column_letter(max(num_cols, 1))
        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = f'Смета №{estimate.number} — {estimate.name}'
        ws['A1'].font = header_font
        ws['A1'].alignment = center

        # Column headers (row 3)
        for col_idx, col_def in enumerate(visible_cols, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_def.get('label', col_def['key']))
            cell.font = bold
            cell.alignment = center
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(col_def.get('width', 100) / 8, 8)

        row_num = 4
        agg_sums = {c['key']: Decimal('0') for c in visible_cols if c.get('aggregatable')}

        for section in sections:
            # Section header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
            cell = ws.cell(row=row_num, column=1, value=section.name)
            cell.font = section_font
            row_num += 1

            section_items = [i for i in items if i.section_id == section.id]
            for item in section_items:
                builtin_values = {
                    'item_number': Decimal(str(item.item_number or 0)),
                    'quantity': item.quantity or Decimal('0'),
                    'material_unit_price': item.material_unit_price or Decimal('0'),
                    'work_unit_price': item.work_unit_price or Decimal('0'),
                    'material_total': item.material_total or Decimal('0'),
                    'work_total': item.work_total or Decimal('0'),
                    'line_total': item.line_total or Decimal('0'),
                }
                custom_data = item.custom_data or {}
                computed = compute_all_formulas(config, builtin_values, custom_data)

                for col_idx, col_def in enumerate(visible_cols, 1):
                    key = col_def['key']
                    col_type = col_def.get('type', 'builtin')
                    value = None

                    if col_type == 'builtin':
                        field = col_def.get('builtin_field', key)
                        value = getattr(item, field, None)
                    elif col_type == 'formula':
                        value = computed.get(key)
                    elif col_type.startswith('custom_'):
                        value = custom_data.get(key, '')

                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.border = thin_border

                    if isinstance(value, Decimal):
                        cell.value = float(value)
                        cell.number_format = num_fmt
                    elif value is not None:
                        try:
                            cell.value = float(value)
                            cell.number_format = num_fmt
                        except (ValueError, TypeError):
                            cell.value = str(value) if value else ''
                    else:
                        cell.value = ''

                    # Accumulate aggregatables
                    if key in agg_sums and value is not None:
                        try:
                            agg_sums[key] += Decimal(str(value))
                        except Exception:
                            pass

                row_num += 1

        # Footer totals
        if agg_sums:
            row_num += 1
            for col_idx, col_def in enumerate(visible_cols, 1):
                key = col_def['key']
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = bold
                cell.border = thin_border
                if key in agg_sums:
                    cell.value = float(agg_sums[key])
                    cell.number_format = num_fmt
                elif col_idx == 1:
                    cell.value = 'ИТОГО'

        # Response
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'Смета_{estimate.number}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class EstimateSectionViewSet(viewsets.ModelViewSet):
    """ViewSet для разделов сметы"""

    queryset = EstimateSection.objects.select_related('estimate').prefetch_related('subsections')
    serializer_class = EstimateSectionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['estimate']

    @action(detail=True, methods=['post'], url_path='demote-to-item')
    def demote_to_item(self, request, pk=None):
        """Превратить раздел обратно в обычную строку сметы."""
        try:
            section = EstimateSection.objects.get(pk=pk)
        except EstimateSection.DoesNotExist:
            return Response({'error': 'Раздел не найден'}, status=status.HTTP_404_NOT_FOUND)

        from .services.estimate_import_service import EstimateImportService
        result = EstimateImportService().demote_section_to_item(int(pk))
        return Response(result)


class EstimateSubsectionViewSet(viewsets.ModelViewSet):
    """ViewSet для подразделов сметы"""
    
    queryset = EstimateSubsection.objects.select_related('section', 'section__estimate')
    serializer_class = EstimateSubsectionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['section']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Фильтрация по смете через раздел
        estimate_id = self.request.query_params.get('estimate')
        if estimate_id:
            queryset = queryset.filter(section__estimate_id=estimate_id)
        return queryset


class EstimateCharacteristicViewSet(viewsets.ModelViewSet):
    """ViewSet для характеристик сметы"""
    
    queryset = EstimateCharacteristic.objects.select_related('estimate')
    serializer_class = EstimateCharacteristicSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['estimate']


class EstimateItemPagination(PageNumberPagination):
    """Пагинация строк сметы. page_size=all отключает пагинацию (обратная совместимость)."""
    page_size = 200
    page_size_query_param = 'page_size'
    max_page_size = 2000

    def paginate_queryset(self, queryset, request, view=None):
        if request.query_params.get('page_size') == 'all':
            return None
        return super().paginate_queryset(queryset, request, view)


class EstimateItemViewSet(viewsets.ModelViewSet):
    """ViewSet для строк сметы"""
    pagination_class = EstimateItemPagination

    queryset = EstimateItem.objects.select_related(
        'estimate', 'section', 'subsection',
        'product', 'work_item', 'source_price_history',
    )
    serializer_class = EstimateItemSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'estimate', 'section', 'subsection', 'product',
        'work_item', 'is_analog',
    ]
    search_fields = ['name', 'model_name', 'original_name']
    ordering_fields = ['sort_order', 'item_number', 'name', 'material_unit_price', 'work_unit_price']
    ordering = ['sort_order', 'item_number']

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        estimate_id = self.request.query_params.get('estimate')
        if estimate_id:
            try:
                estimate = Estimate.objects.only('column_config').get(pk=estimate_id)
                ctx['column_config'] = estimate.column_config or []
            except Estimate.DoesNotExist:
                pass
        return ctx

    def get_queryset(self):
        return super().get_queryset().annotate(
            _material_total=ExpressionWrapper(
                F('quantity') * F('material_unit_price'),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            _work_total=ExpressionWrapper(
                F('quantity') * F('work_unit_price'),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            _line_total=ExpressionWrapper(
                F('quantity') * F('material_unit_price') + F('quantity') * F('work_unit_price'),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )

    @action(detail=True, methods=['post'], url_path='promote-to-section')
    def promote_to_section(self, request, pk=None):
        """Превратить строку сметы в раздел (секцию)."""
        try:
            EstimateItem.objects.get(pk=pk)
        except EstimateItem.DoesNotExist:
            return Response({'error': 'Строка не найдена'}, status=status.HTTP_404_NOT_FOUND)

        from .services.estimate_import_service import EstimateImportService
        result = EstimateImportService().promote_item_to_section(int(pk))
        return Response(result)

    @action(detail=True, methods=['post'], url_path='move')
    def move(self, request, pk=None):
        """Переместить строку сметы вверх/вниз или в другой раздел."""
        try:
            EstimateItem.objects.get(pk=pk)
        except EstimateItem.DoesNotExist:
            return Response({'error': 'Строка не найдена'}, status=status.HTTP_404_NOT_FOUND)

        from .services.estimate_import_service import EstimateImportService
        service = EstimateImportService()

        direction = request.data.get('direction')
        target_section_id = request.data.get('target_section_id')

        if direction in ('up', 'down'):
            if direction == 'up':
                result = service.move_item_up(int(pk))
            else:
                result = service.move_item_down(int(pk))
        elif target_section_id is not None:
            try:
                EstimateSection.objects.get(pk=target_section_id)
            except EstimateSection.DoesNotExist:
                return Response({'error': 'Раздел не найден'}, status=status.HTTP_404_NOT_FOUND)
            result = service.move_item_to_section(int(pk), int(target_section_id))
        else:
            return Response(
                {'error': 'Укажите direction (up/down) или target_section_id'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(result)

    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """Создать множество строк сметы за одну операцию"""
        serializer = EstimateItemBulkCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        items = serializer.save()
        return Response(
            EstimateItemSerializer(items, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='bulk-move')
    def bulk_move(self, request):
        """Переместить группу строк на указанную позицию."""
        item_ids = request.data.get('item_ids', [])
        target_position = request.data.get('target_position')

        if not item_ids or not isinstance(item_ids, list):
            return Response(
                {'error': 'Необходим непустой массив item_ids'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if target_position is None or not isinstance(target_position, int) or target_position < 1:
            return Response(
                {'error': 'target_position должен быть целым числом >= 1'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .services.estimate_import_service import EstimateImportService
        result = EstimateImportService().bulk_move_items(item_ids, target_position)
        return Response(result)

    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update_items(self, request):
        """Обновить множество строк сметы за одну операцию.
        Ожидает массив объектов с обязательным полем 'id'."""
        items_data = request.data
        if not isinstance(items_data, list):
            return Response(
                {'error': 'Ожидается массив объектов'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ids = [item.get('id') for item in items_data if item.get('id')]
        existing = {item.id: item for item in EstimateItem.objects.filter(id__in=ids)}
        updated = []

        for item_data in items_data:
            item_id = item_data.get('id')
            if not item_id or item_id not in existing:
                continue
            obj = existing[item_id]
            for field, value in item_data.items():
                if field == 'id':
                    continue
                setattr(obj, field, value)
            updated.append(obj)

        if updated:
            update_fields = [
                k for k in items_data[0].keys() if k != 'id'
            ]
            EstimateItem.objects.bulk_update(updated, update_fields)

        return Response(
            EstimateItemSerializer(updated, many=True).data,
        )

    @action(detail=False, methods=['post'], url_path='auto-match')
    def auto_match(self, request):
        """Автоматический подбор цен и работ для строк сметы"""
        estimate_id = request.data.get('estimate_id')
        price_list_id = request.data.get('price_list_id')

        if not estimate_id:
            return Response(
                {'error': 'Не указан estimate_id'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            estimate = Estimate.objects.get(pk=estimate_id)
        except Estimate.DoesNotExist:
            return Response(
                {'error': 'Смета не найдена'},
                status=status.HTTP_404_NOT_FOUND,
            )

        supplier_ids = request.data.get('supplier_ids', [])
        price_strategy = request.data.get('price_strategy', 'cheapest')

        from .services.estimate_auto_matcher import EstimateAutoMatcher
        matcher = EstimateAutoMatcher()
        results = matcher.preview_matches(
            estimate,
            supplier_ids=supplier_ids or None,
            price_strategy=price_strategy,
        )
        return Response(results)

    @action(detail=False, methods=['post'], url_path='auto-match-works')
    def auto_match_works(self, request):
        """Preview-подбор работ для строк сметы (без сохранения)"""
        estimate_id = request.data.get('estimate_id')
        price_list_id = request.data.get('price_list_id')

        if not estimate_id:
            return Response(
                {'error': 'Не указан estimate_id'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            estimate = Estimate.objects.get(pk=estimate_id)
        except Estimate.DoesNotExist:
            return Response(
                {'error': 'Смета не найдена'},
                status=status.HTTP_404_NOT_FOUND,
            )

        from .services.estimate_auto_matcher import EstimateAutoMatcher
        matcher = EstimateAutoMatcher()
        results = matcher.preview_works(estimate, price_list_id=price_list_id)
        return Response(results)

    @action(detail=False, methods=['post'], url_path='apply-match-works')
    def apply_match_works(self, request):
        """Применить выбранные совпадения работ и записать маппинги"""
        items_data = request.data.get('items', [])
        if not items_data:
            return Response(
                {'error': 'Пустой список'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from pricelists.models import WorkItem
        from .services.estimate_auto_matcher import EstimateAutoMatcher

        ids = [d['item_id'] for d in items_data if d.get('item_id')]
        existing = {
            item.id: item
            for item in EstimateItem.objects.filter(id__in=ids).select_related('product')
        }

        work_ids = [d['work_item_id'] for d in items_data if d.get('work_item_id')]
        works = {wi.id: wi for wi in WorkItem.objects.filter(id__in=work_ids)}

        updated = []
        for d in items_data:
            item = existing.get(d.get('item_id'))
            wi = works.get(d.get('work_item_id'))
            if not item or not wi:
                continue

            item.work_item = wi
            if d.get('work_price') and item.work_unit_price == 0:
                item.work_unit_price = d['work_price']
            updated.append(item)

            # Записываем маппинг для обучения системы
            if item.product:
                EstimateAutoMatcher.record_manual_correction(item.product, wi)

        if updated:
            EstimateItem.objects.bulk_update(updated, ['work_item', 'work_unit_price'])

        return Response({'applied': len(updated)})

    @action(detail=False, methods=['post'], url_path='import',
            parser_classes=[MultiPartParser])
    def import_file(self, request):
        """Импорт строк сметы из Excel или PDF.
        Если preview=true, возвращает предпросмотр без сохранения."""
        estimate_id = request.data.get('estimate_id')
        file = request.FILES.get('file')
        preview_mode = request.data.get('preview', '').lower() in ('true', '1')

        if not estimate_id or not file:
            return Response(
                {'error': 'Необходимы estimate_id и file'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            estimate = Estimate.objects.get(pk=estimate_id)
        except Estimate.DoesNotExist:
            return Response(
                {'error': 'Смета не найдена'},
                status=status.HTTP_404_NOT_FOUND,
            )

        file_content = file.read()
        filename = file.name.lower()

        from .services.estimate_import_service import EstimateImportService
        importer = EstimateImportService()

        if filename.endswith(('.xlsx', '.xls')):
            parsed = importer.import_from_excel(file_content, file.name)
        elif filename.endswith('.pdf'):
            parsed = importer.import_from_pdf(file_content, file.name)
        else:
            return Response(
                {'error': 'Поддерживаются только файлы Excel (.xlsx) и PDF'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if preview_mode:
            return Response({
                'rows': [row.model_dump() for row in parsed.rows],
                'sections': parsed.sections,
                'total_rows': parsed.total_rows,
                'confidence': parsed.confidence,
            })

        created_items = importer.save_imported_items(int(estimate_id), parsed)
        return Response(
            EstimateItemSerializer(created_items, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='import-rows')
    def import_rows(self, request):
        """Импорт строк из предпросмотра (JSON) с назначенными разделами."""
        estimate_id = request.data.get('estimate_id')
        rows = request.data.get('rows', [])

        if not estimate_id or not rows:
            return Response(
                {'error': 'Необходимы estimate_id и rows'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            Estimate.objects.get(pk=estimate_id)
        except Estimate.DoesNotExist:
            return Response(
                {'error': 'Смета не найдена'},
                status=status.HTTP_404_NOT_FOUND,
            )

        from .services.estimate_import_service import EstimateImportService
        importer = EstimateImportService()
        created_items = importer.save_rows_from_preview(int(estimate_id), rows)
        return Response(
            EstimateItemSerializer(created_items, many=True).data,
            status=status.HTTP_201_CREATED,
        )

    # ── Постраничный импорт PDF ──────────────────────────────────

    @action(detail=False, methods=['post'], url_path='import-pdf',
            parser_classes=[MultiPartParser])
    def import_pdf(self, request):
        """Запуск постраничного импорта PDF. Возвращает session_id сразу."""
        estimate_id = request.data.get('estimate_id')
        file = request.FILES.get('file')

        if not estimate_id or not file:
            return Response(
                {'error': 'Необходимы estimate_id и file'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            Estimate.objects.get(pk=estimate_id)
        except Estimate.DoesNotExist:
            return Response(
                {'error': 'Смета не найдена'},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not file.name.lower().endswith('.pdf'):
            return Response(
                {'error': 'Только PDF файлы'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        file_content = file.read()

        from .tasks import create_import_session, process_estimate_pdf_pages
        session = create_import_session(file_content, int(estimate_id))
        process_estimate_pdf_pages.delay(session['session_id'])

        return Response(session, status=status.HTTP_202_ACCEPTED)

    @action(detail=False, methods=['get'],
            url_path=r'import-progress/(?P<session_id>[a-f0-9]+)')
    def import_progress(self, request, session_id=None):
        """Поллинг прогресса импорта PDF."""
        from .tasks import get_session_data
        data = get_session_data(session_id)
        if not data:
            return Response(
                {'error': 'Сессия не найдена или истекла'},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(data)

    @action(detail=False, methods=['post'],
            url_path=r'import-cancel/(?P<session_id>[a-f0-9]+)')
    def import_cancel(self, request, session_id=None):
        """Отмена импорта PDF."""
        from .tasks import cancel_session
        if cancel_session(session_id):
            return Response({'status': 'cancelled'})
        return Response(
            {'error': 'Сессия не найдена'},
            status=status.HTTP_404_NOT_FOUND,
        )


class MountingEstimateViewSet(viewsets.ModelViewSet):
    """ViewSet для монтажных смет"""
    
    queryset = MountingEstimate.objects.select_related(
        'object', 'source_estimate', 'agreed_counterparty', 
        'created_by', 'parent_version'
    )
    serializer_class = MountingEstimateSerializer
    
    def perform_create(self, serializer):
        """Автоматически устанавливаем created_by из текущего пользователя"""
        serializer.save(created_by=self.request.user)
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = [
        'object', 'source_estimate', 'status', 'agreed_counterparty'
    ]
    search_fields = ['number', 'name']
    
    @action(detail=False, methods=['post'], url_path='from-estimate')
    def from_estimate(self, request):
        """Создать монтажную смету из обычной сметы"""
        serializer = MountingEstimateCreateFromEstimateSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        mounting_estimate = serializer.save()
        return Response(
            MountingEstimateSerializer(mounting_estimate).data,
            status=status.HTTP_201_CREATED
        )
    
    @action(detail=True, methods=['post'], url_path='create-version')
    def create_version(self, request, pk=None):
        """Создать новую версию монтажной сметы"""
        mounting_estimate = self.get_object()
        new_version = mounting_estimate.create_new_version()
        serializer = MountingEstimateSerializer(new_version)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def agree(self, request, pk=None):
        """Согласовать с Исполнителем"""
        mounting_estimate = self.get_object()
        counterparty_id = request.data.get('counterparty_id')
        
        if not counterparty_id:
            return Response(
                {'error': 'Не указан ID контрагента'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from accounting.models import Counterparty
        try:
            counterparty = Counterparty.objects.get(id=counterparty_id)
            if counterparty.type not in [Counterparty.Type.VENDOR, Counterparty.Type.BOTH]:
                return Response(
                    {'error': 'Контрагент должен быть типа "Исполнитель/Поставщик" или "Заказчик и Исполнитель"'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            mounting_estimate.agreed_counterparty = counterparty
            mounting_estimate.agreed_date = timezone.now().date()
            mounting_estimate.status = MountingEstimate.Status.APPROVED
            mounting_estimate.save()
            serializer = MountingEstimateSerializer(mounting_estimate)
            return Response(serializer.data)
        except Counterparty.DoesNotExist:
            return Response(
                {'error': 'Контрагент не найден'},
                status=status.HTTP_404_NOT_FOUND
            )


class ColumnConfigTemplateViewSet(viewsets.ModelViewSet):
    """ViewSet для шаблонов конфигурации столбцов."""

    queryset = ColumnConfigTemplate.objects.select_related('created_by')
    serializer_class = ColumnConfigTemplateSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='apply')
    def apply_to_estimate(self, request, pk=None):
        """Применить шаблон к смете (копирует column_config)."""
        template = self.get_object()
        estimate_id = request.data.get('estimate_id')
        if not estimate_id:
            return Response(
                {'error': 'Не указан estimate_id'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            estimate = Estimate.objects.get(pk=estimate_id)
        except Estimate.DoesNotExist:
            return Response(
                {'error': 'Смета не найдена'},
                status=status.HTTP_404_NOT_FOUND,
            )
        estimate.column_config = template.column_config
        estimate.save(update_fields=['column_config'])
        return Response({'status': 'ok', 'estimate_id': estimate.id})
