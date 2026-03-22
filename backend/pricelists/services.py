"""
pricelists/services.py — бизнес-логика для прайс-листов.

Вынесена из views.py для тонких view-обёрток.
"""
from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Управление позициями прайс-листа
# ---------------------------------------------------------------------------

def add_items_to_pricelist(price_list, work_item_ids):
    """
    Добавить работы в прайс-лист.
    Возвращает dict {'added': [...], 'count': int}.
    """
    from .models import PriceListItem, WorkItem

    work_items = WorkItem.objects.filter(id__in=work_item_ids, is_current=True)
    work_items_dict = {wi.id: wi for wi in work_items}

    existing_items = PriceListItem.objects.filter(
        price_list=price_list,
        work_item__in=work_items,
    ).select_related('work_item')
    existing_work_item_ids = {item.work_item_id for item in existing_items}

    added = []

    # Обновляем неактивные элементы через bulk_update
    items_to_update = [item for item in existing_items if not item.is_included]
    if items_to_update:
        for item in items_to_update:
            item.is_included = True
            added.append(item.work_item_id)
        PriceListItem.objects.bulk_update(items_to_update, ['is_included'])

    # Создаём новые элементы через bulk_create
    new_work_items = [
        work_items_dict[wid]
        for wid in work_item_ids
        if wid in work_items_dict and wid not in existing_work_item_ids
    ]
    if new_work_items:
        new_items = [
            PriceListItem(price_list=price_list, work_item=wi, is_included=True)
            for wi in new_work_items
        ]
        PriceListItem.objects.bulk_create(new_items)
        added.extend([wi.id for wi in new_work_items])

    return {'added': added, 'count': len(added)}


def remove_items_from_pricelist(price_list, work_item_ids):
    """
    Удалить работы из прайс-листа.
    Возвращает dict {'removed': [...], 'count': int}.
    """
    from .models import PriceListItem

    deleted_count = PriceListItem.objects.filter(
        price_list=price_list,
        work_item_id__in=work_item_ids,
    ).delete()[0]

    return {'removed': work_item_ids, 'count': deleted_count}


# ---------------------------------------------------------------------------
# Версионирование работ
# ---------------------------------------------------------------------------

def create_work_item_version(instance, update_data):
    """
    Создать новую версию WorkItem и применить изменения.
    Возвращает новую версию.
    """
    new_version = instance.create_new_version()

    for attr, value in update_data.items():
        if attr not in ('article', 'version_number', 'is_current', 'parent_version'):
            setattr(new_version, attr, value)
    new_version.save()

    return new_version


# ---------------------------------------------------------------------------
# Экспорт прайс-листа в Excel
# ---------------------------------------------------------------------------

def export_pricelist_to_excel(price_list):
    """
    Сформировать Excel-файл прайс-листа.
    Возвращает (bytes_content, filename).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    # Стили
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Заголовок
    ws.merge_cells('A1:I1')
    ws['A1'] = f"Прайс-лист №{price_list.number} от {price_list.date.strftime('%d.%m.%Y')}"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align

    # Название (если есть)
    if price_list.name:
        ws.merge_cells('A2:I2')
        ws['A2'] = price_list.name
        ws['A2'].alignment = center_align

    # Ставки по разрядам
    row = 4
    ws[f'A{row}'] = "Ставки по разрядам:"
    ws[f'A{row}'].font = bold_font
    row += 1

    for grade_num in range(1, 6):
        rate = price_list.get_rate_for_grade(grade_num)
        ws[f'A{row}'] = f"Разряд {grade_num}:"
        ws[f'B{row}'] = f"{rate} руб/ч"
        row += 1

    # Пустая строка
    row += 1

    # Заголовки таблицы работ
    ws[f'A{row}'] = "Работы:"
    ws[f'A{row}'].font = bold_font
    row += 1

    headers = [
        'Артикул', 'Раздел', 'Наименование', 'Ед.изм.',
        'Часы', 'Разряд', 'Коэфф.', 'Стоимость', 'Комментарий',
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Устанавливаем ширину столбцов
    col_widths = {'A': 12, 'B': 15, 'C': 40, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 15, 'I': 30}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    row += 1

    # Данные работ
    items_qs = price_list.items.filter(is_included=True).select_related(
        'work_item', 'work_item__section', 'work_item__grade',
    )
    for item in items_qs:
        work = item.work_item
        effective_grade = float(item.effective_grade)
        values = [
            work.article,
            work.section.code,
            work.name,
            work.unit,
            float(item.effective_hours),
            effective_grade,
            float(item.effective_coefficient),
            float(item.calculated_cost),
            work.comment if work.comment else '',
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in (5, 6, 7, 8):
                cell.alignment = center_align
            elif col == 9:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        row += 1

    # Итого
    total = sum(
        item.calculated_cost
        for item in price_list.items.filter(is_included=True)
    )
    row += 1
    ws[f'G{row}'] = "ИТОГО:"
    ws[f'G{row}'].font = bold_font
    ws[f'H{row}'] = float(total)
    ws[f'H{row}'].font = bold_font
    ws.merge_cells(f'I{row}:I{row}')

    # Сохраняем в bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"pricelist_{price_list.number}_{price_list.date.strftime('%Y%m%d')}.xlsx"
    return output.read(), filename
